from __future__ import annotations

import csv
import io
import re
import unicodedata
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import requests

from .config import get_settings

settings = get_settings()

MAX_FILE_TEXT = 15000
MAX_TOTAL_TEXT = 60000


def _plain(value: Any) -> str:
    return str(value or "").strip()


def _normalize(value: str) -> str:
    value = unicodedata.normalize("NFKD", value)
    value = "".join(ch for ch in value if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", value.lower()).strip()


def _snippet(original: str, needles: list[str], width: int = 260) -> str:
    normalized = _normalize(original)
    positions = [normalized.find(_normalize(term)) for term in needles if _normalize(term) in normalized]
    position = min(positions) if positions else 0
    start = max(0, position - 70)
    end = min(len(original), start + width)
    result = original[start:end].strip()
    return f"…{result}…" if start or end < len(original) else result


def _storage_bytes(storage_path: str) -> bytes:
    response = requests.get(
        f"{settings.supabase_url.rstrip('/')}/storage/v1/object/diagnose-evidence/{storage_path}",
        headers={
            "apikey": settings.supabase_service_role_key,
            "Authorization": f"Bearer {settings.supabase_service_role_key}",
        },
        timeout=30,
    )
    if response.status_code != 200:
        return b""
    return response.content


def _extract_file_text(item: dict[str, Any]) -> tuple[str, str | None]:
    path = _plain(item.get("storage_path"))
    if not path:
        return "", None
    body = _storage_bytes(path)
    if not body:
        return "", "No se pudo leer el archivo almacenado."

    mime = _plain(item.get("mime_type")).lower()
    suffix = Path(path).suffix.lower()
    try:
        if mime.startswith("text/") or suffix in {".txt", ".csv"}:
            decoded = body.decode("utf-8", errors="replace")
            if suffix == ".csv" or "csv" in mime:
                rows = csv.reader(io.StringIO(decoded))
                decoded = "\n".join(" | ".join(cell for cell in row) for row in list(rows)[:250])
            return decoded[:MAX_FILE_TEXT], None

        if mime == "application/pdf" or suffix == ".pdf":
            from pypdf import PdfReader

            reader = PdfReader(io.BytesIO(body))
            text = "\n".join((page.extract_text() or "") for page in reader.pages[:30])
            return text[:MAX_FILE_TEXT], None if text.strip() else "El PDF no contiene texto seleccionable; agrega una nota descriptiva."

        if suffix == ".docx" or "wordprocessingml" in mime:
            from docx import Document

            document = Document(io.BytesIO(body))
            text = "\n".join(paragraph.text for paragraph in document.paragraphs)
            return text[:MAX_FILE_TEXT], None

        if suffix == ".xlsx" or "spreadsheetml" in mime:
            from openpyxl import load_workbook

            workbook = load_workbook(io.BytesIO(body), read_only=True, data_only=True)
            lines: list[str] = []
            for sheet in workbook.worksheets[:8]:
                lines.append(f"Hoja: {sheet.title}")
                for row in sheet.iter_rows(max_row=250, values_only=True):
                    lines.append(" | ".join(_plain(cell) for cell in row if cell is not None))
            return "\n".join(lines)[:MAX_FILE_TEXT], None

        if mime.startswith("image/") or suffix in {".png", ".jpg", ".jpeg", ".webp"}:
            return "", "Las capturas se analizan mediante el nombre y las notas. Describe el dato visible más importante."
    except Exception:
        return "", "El archivo no pudo convertirse a texto; agrega una nota con el hallazgo principal."

    return "", "Este formato se conserva como evidencia, pero no tiene extracción automática de texto."


def build_evidence_corpus(
    diagnosis: dict[str, Any],
    evidence: list[dict[str, Any]],
    interview_questions: list[dict[str, Any]],
) -> tuple[str, list[dict[str, Any]], list[str]]:
    blocks = [
        f"Empresa: {_plain(diagnosis.get('company_name'))}",
        f"Industria: {_plain(diagnosis.get('industry'))}",
        f"Objetivo: {_plain(diagnosis.get('objective'))}",
        f"Problema declarado: {_plain(diagnosis.get('declared_problem'))}",
        f"Resumen ejecutivo: {_plain(diagnosis.get('executive_summary'))}",
    ]
    evidence_context: list[dict[str, Any]] = []
    limitations: list[str] = []

    for item in evidence:
        extracted, limitation = _extract_file_text(item) if item.get("storage_path") else ("", None)
        block = "\n".join(filter(None, [
            f"Evidencia: {_plain(item.get('name'))}",
            f"Categoría: {_plain(item.get('category'))}",
            f"Notas: {_plain(item.get('notes'))}",
            f"Contenido extraído: {extracted}",
            f"Enlace: {_plain(item.get('external_url'))}",
        ]))
        blocks.append(block)
        evidence_context.append({
            "id": item.get("id"),
            "name": item.get("name"),
            "category": item.get("category"),
            "has_notes": bool(_plain(item.get("notes"))),
            "extracted_chars": len(extracted),
            "limitation": limitation,
        })
        if limitation and limitation not in limitations:
            limitations.append(limitation)

    answered = [q for q in interview_questions if _plain(q.get("answer")) and q.get("status") == "answered"]
    if answered:
        blocks.append("Respuestas de entrevista:\n" + "\n".join(
            f"- {_plain(row.get('question'))}: {_plain(row.get('answer'))}" for row in answered
        ))

    corpus = "\n\n".join(blocks)[:MAX_TOTAL_TEXT]
    return corpus, evidence_context, limitations


def analyze_corpus(
    diagnosis: dict[str, Any],
    evidence: list[dict[str, Any]],
    interview_questions: list[dict[str, Any]],
) -> dict[str, Any]:
    corpus, evidence_context, limitations = build_evidence_corpus(diagnosis, evidence, interview_questions)
    normalized = _normalize(corpus)
    signals: list[dict[str, Any]] = []
    suggestions: list[dict[str, Any]] = []

    def contains(*phrases: str) -> bool:
        return any(_normalize(phrase) in normalized for phrase in phrases)

    def add_signal(
        key: str,
        kind: str,
        section: str,
        title: str,
        description: str,
        terms: list[str],
        impact: str,
        urgency: str,
        recommendation: str,
        confidence: int,
        assessment: tuple[str, str, int, str] | None = None,
    ) -> None:
        if any(item["key"] == key for item in signals):
            return
        signals.append({
            "key": key,
            "kind": kind,
            "section": section,
            "title": title,
            "description": description,
            "evidence": _snippet(corpus, terms),
            "impact": impact,
            "urgency": urgency,
            "recommendation": recommendation,
            "confidence": confidence,
        })
        if assessment:
            a_section, question_id, score, rationale = assessment
            suggestions.append({
                "section": a_section,
                "question_id": question_id,
                "suggested_score": score,
                "rationale": rationale,
                "confidence": confidence,
            })

    hour_matches = [float(value.replace(",", ".")) for value in re.findall(r"(\d+(?:[.,]\d+)?)\s*(?:horas?|hrs?|h)\b", normalized)]
    delay_hours = max(hour_matches) if hour_matches else 0
    delay_terms = ["demora", "tardó", "tardaron", "horas sin respuesta", "respondieron a las", "tiempo de respuesta"]
    if delay_hours >= 1 or contains(*delay_terms):
        critical = delay_hours >= 4
        add_signal(
            "response-delay", "risk", "conversion", "Tiempo de respuesta elevado",
            "La evidencia muestra una demora relevante entre la consulta y la primera respuesta.",
            delay_terms + ([f"{delay_hours:g} horas"] if delay_hours else []),
            "critical" if critical else "high", "high",
            "Definir un SLA por canal, responsables de respaldo y alertas para consultas sin atender.",
            94 if delay_hours else 78,
            ("conversion", "response_time", 0 if critical else 1, "La evidencia indica una respuesta tardía o sin estándar operativo."),
        )

    if contains("sin seguimiento", "no existe seguimiento", "no se observa seguimiento", "no hubo seguimiento", "sin segundo seguimiento", "no volvió a contactar", "ni seguimiento posterior", "sin seguimiento posterior"):
        add_signal(
            "missing-followup", "risk", "conversion", "Seguimiento comercial no estructurado",
            "No se evidencia una cadencia clara para recuperar oportunidades que no reservan o responden en el primer contacto.",
            ["sin seguimiento", "no existe seguimiento", "no se observa seguimiento", "no hubo seguimiento", "ni seguimiento posterior"],
            "critical", "high",
            "Crear una cadencia mínima con intentos, canales, guiones, responsable y próxima fecha.",
            92,
            ("conversion", "followup_process", 0, "La evidencia indica ausencia de una cadencia de seguimiento."),
        )

    if contains("sin crm", "no existe crm", "no se detectó crm", "no se detecto crm", "directamente desde whatsapp", "solo whatsapp"):
        add_signal(
            "no-central-crm", "risk", "automation", "Ausencia de una base comercial central",
            "La operación parece depender de canales o registros aislados, sin una fuente única para estados, responsables y próximos pasos.",
            ["sin crm", "no existe crm", "directamente desde whatsapp", "solo whatsapp"],
            "high", "high",
            "Centralizar leads, actividad, responsables, estados y seguimientos en una base operativa única.",
            88,
            ("automation", "central_crm", 0, "No se encontró evidencia de una base central confiable."),
        )

    if contains("manual", "copiar y pegar", "hoja de cálculo", "hoja de calculo", "excel", "depende de la memoria"):
        add_signal(
            "manual-operation", "risk", "process", "Dependencia de tareas manuales",
            "La evidencia sugiere retrabajo, coordinación manual o dependencia de la memoria del equipo.",
            ["manual", "copiar y pegar", "excel", "depende de la memoria"],
            "medium", "medium",
            "Mapear las tareas repetitivas, eliminar duplicación y automatizar solamente los pasos estables.",
            76,
            ("process", "manual_rework", 1, "Se detectaron señales de retrabajo o coordinación manual."),
        )

    if contains("sin métricas", "sin metricas", "no se mide", "no mide", "no existe medición", "no existe medicion", "no conocemos la conversión", "no conocemos la conversion"):
        add_signal(
            "missing-metrics", "risk", "conversion", "Conversión sin medición periódica",
            "No se encontraron indicadores confiables para saber dónde se pierden oportunidades o qué canal convierte mejor.",
            ["sin métricas", "no se mide", "no existe medición", "no conocemos la conversión"],
            "high", "medium",
            "Medir volumen, tiempo de respuesta, contacto, cita, propuesta, venta e ingreso por fuente.",
            86,
            ("conversion", "conversion_metrics", 0, "La evidencia indica ausencia de métricas de conversión."),
        )

    if contains("sin guion", "no hay guion", "improvis", "sin script", "no existen scripts"):
        add_signal(
            "missing-script", "risk", "conversion", "Primer contacto dependiente de improvisación",
            "El equipo no parece utilizar un guion consistente para calificar, comunicar valor y avanzar al siguiente paso.",
            ["sin guion", "no hay guion", "improvis", "sin script"],
            "medium", "medium",
            "Crear un guion breve por servicio con apertura, preguntas, propuesta de valor y CTA.",
            84,
            ("conversion", "first_contact_script", 0, "La evidencia señala ausencia de un guion estandarizado."),
        )

    if contains("sin invitación", "sin invitacion", "no se observa una invitación", "no se observa una invitacion", "no se observa una pregunta", "cta débil", "cta debil"):
        add_signal(
            "weak-booking-cta", "risk", "conversion", "CTA de agendamiento débil",
            "La conversación no conduce de forma clara a una cita, reserva o siguiente paso operativo.",
            ["sin invitación", "no se observa una invitación", "no se observa una pregunta", "reserv", "cta débil"],
            "high", "high",
            "Cerrar cada conversación con dos opciones concretas de horario o una acción verificable.",
            82,
            ("conversion", "appointment_process", 1, "No se observa una invitación clara a reservar o confirmar."),
        )

    if contains("información dispersa", "informacion dispersa", "entre chats", "varias hojas", "diferentes personas", "se pierde información", "se pierde informacion"):
        add_signal(
            "fragmented-information", "risk", "process", "Información fragmentada",
            "Los datos y el contexto del cliente parecen repartidos entre herramientas, conversaciones o personas.",
            ["información dispersa", "entre chats", "varias hojas", "se pierde información"],
            "high", "medium",
            "Definir una fuente única y campos obligatorios en cada interacción y entrega.",
            86,
            ("process", "single_source_truth", 0, "La evidencia muestra fragmentación de información."),
        )

    high_value_terms = [term for term in ["implantes", "ortodoncia", "diseño de sonrisa", "diseno de sonrisa", "estética dental", "estetica dental"] if term in normalized]
    if high_value_terms:
        add_signal(
            "high-value-services", "opportunity", "icp", "Servicios de alto valor detectados",
            "La empresa ofrece servicios con potencial de ticket e impacto comercial relevantes.",
            high_value_terms,
            "high", "medium",
            "Priorizar la calificación, los guiones y la medición por servicio de alto valor.",
            90,
            ("icp", "high_value_services", 3, "La evidencia identifica servicios de mayor valor."),
        )

    if not evidence:
        limitations.append("No hay evidencias guardadas; el análisis se basa únicamente en el contexto inicial del diagnóstico.")
    if any(item.get("limitation") and "capturas" in item["limitation"].lower() for item in evidence_context):
        limitations.append("Aura no interpreta píxeles en esta versión. Para capturas, utiliza las notas para describir horas, mensajes y hechos visibles.")
    limitations.append("El análisis es un borrador asistido por reglas y debe ser validado por Laura antes de modificar evaluaciones o hallazgos.")
    limitations = list(dict.fromkeys(limitations))

    answered_keys = {str(q.get("question_key")) for q in interview_questions if q.get("status") == "answered" and _plain(q.get("answer"))}
    question_bank = [
        ("lead-volume", "conversion", "¿Cuántas consultas nuevas reciben al mes y cómo se distribuyen por WhatsApp, Instagram, web, llamadas y referidos?", "Permite dimensionar la demanda real y priorizar los canales.", "high", ["consultas al mes", "leads al mes"]),
        ("response-time", "conversion", "¿Cuál es el tiempo promedio y el tiempo máximo de primera respuesta durante horario laboral?", "Confirma si la demora observada es un caso aislado o una fuga sistemática.", "critical" if any(s["key"] == "response-delay" for s in signals) else "high", ["tiempo promedio de respuesta", "sla de respuesta"]),
        ("conversion-funnel", "conversion", "De cada 100 consultas, ¿cuántas reciben respuesta, cuántas agendan, cuántas asisten y cuántas compran?", "Construye la línea base del embudo y localiza la mayor fuga.", "critical", ["de cada 100", "tasa de conversión"]),
        ("followup-cadence", "conversion", "¿Qué ocurre cuando una persona no responde o no reserva en el primer contacto? Indica número de intentos, días, canales y guiones.", "Valida la existencia y calidad de la cadencia de seguimiento.", "critical" if any(s["key"] == "missing-followup" for s in signals) else "high", ["seguimiento a las 24", "seguimiento a las 72", "cadencia"]),
        ("no-show", "conversion", "¿Qué porcentaje de citas no se presenta y qué proceso usan para confirmar o recuperar esos no-shows?", "Los no-shows pueden ocultar una fuga distinta al agendamiento.", "high", ["no-show", "no show", "no se presentan"]),
        ("objections", "conversion", "¿Cuáles son las cinco objeciones más frecuentes y cuáles suelen terminar en pérdida de la oportunidad?", "Convierte conversaciones repetidas en scripts y mejoras de oferta.", "medium", ["objeciones más frecuentes", "objeciones frecuentes"]),
        ("best-customer", "icp", "¿Qué tipo de paciente o cliente produce mejores resultados, mayor margen y menor fricción comercial?", "Define el ICP usando evidencia económica y operativa.", "high", ["cliente ideal", "paciente ideal"]),
        ("service-economics", "icp", "¿Cuáles servicios generan mayor ingreso y margen, cuál es su ticket promedio y cuánto dura su ciclo de decisión?", "Prioriza los servicios que justifican inversión en captación y seguimiento.", "high", ["ticket promedio", "mayor margen"]),
        ("decision-authority", "icp", "¿Quién decide cambios en marketing, ventas, atención y tecnología, y quién puede bloquearlos?", "Identifica al decisor real y la ruta de aprobación.", "high", ["quién decide", "quien decide", "decisor"]),
        ("lead-owner", "process", "¿Quién es responsable de cada lead desde que entra hasta que compra o se descarta?", "Aclara ownership, handoffs y responsabilidad por resultados.", "high", ["responsable de cada lead", "ownership"]),
        ("source-of-truth", "process", "¿Dónde se guarda hoy el historial completo del lead y cómo sabe el equipo cuál es el próximo paso?", "Confirma si existe una fuente única y si el proceso depende de memoria o chats.", "critical" if any(s["key"] in {"no-central-crm", "fragmented-information"} for s in signals) else "high", ["fuente única", "fuente unica", "historial completo"]),
        ("sops-handoffs", "process", "¿Qué partes del proceso están documentadas y qué información debe acompañar una entrega entre personas?", "Detecta dependencia de conocimiento informal y pérdida de contexto.", "medium", ["sop", "proceso documentado"]),
        ("current-tools", "automation", "¿Qué herramientas usan para WhatsApp, agenda, CRM, formularios, reportes y comunicación interna?", "Permite diseñar automatizaciones compatibles con la operación actual.", "high", ["herramientas actuales", "usamos kommo", "usamos hubspot"]),
        ("manual-work", "automation", "¿Qué tareas repetitivas consumen más tiempo cada semana y cuáles requieren copiar datos entre sistemas?", "Prioriza automatizaciones por impacto, frecuencia y riesgo.", "high" if any(s["key"] == "manual-operation" for s in signals) else "medium", ["tareas manuales", "copiar datos"]),
        ("reporting-rhythm", "automation", "¿Qué métricas revisan, con qué frecuencia, quién las prepara y cuánto tarda consolidarlas?", "Valida si la operación aprende de resultados o trabaja sin retroalimentación.", "high" if any(s["key"] == "missing-metrics" for s in signals) else "medium", ["métricas que revisan", "metricas que revisan"]),
        ("ninety-day-outcome", "general", "Si este diagnóstico funciona, ¿qué resultado medible debe haber mejorado dentro de 90 días?", "Alinea el roadmap con un resultado verificable y evita recomendaciones genéricas.", "critical", ["resultado en 90 días", "resultado en 90 dias"]),
    ]

    questions: list[dict[str, Any]] = []
    for key, section, question, rationale, priority, known_patterns in question_bank:
        if key in answered_keys:
            continue
        # Si el corpus contiene una respuesta claramente documentada, la pregunta sigue disponible
        # con menor prioridad para confirmación del decisor.
        known = any(_normalize(pattern) in normalized for pattern in known_patterns)
        questions.append({
            "question_key": key,
            "section": section,
            "question": question,
            "rationale": rationale,
            "priority": "medium" if known and priority in {"high", "critical"} else priority,
        })

    risk_count = sum(1 for item in signals if item["kind"] == "risk")
    opportunity_count = sum(1 for item in signals if item["kind"] == "opportunity")
    top_titles = [item["title"] for item in signals[:3]]
    if signals:
        summary = (
            f"Aura revisó {len(evidence)} evidencias y detectó {risk_count} riesgos"
            f"{f' y {opportunity_count} oportunidades' if opportunity_count else ''}. "
            f"Las señales principales son: {', '.join(top_titles)}. "
            f"Quedan {len(questions)} preguntas sugeridas para validar con el decisor."
        )
    else:
        summary = (
            f"Aura revisó {len(evidence)} evidencias, pero no encontró señales suficientemente explícitas. "
            "Agrega notas con hechos observables, tiempos, herramientas y resultados; luego vuelve a analizar."
        )

    return {
        "engine_version": "aura-local-rules-v1",
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "summary": summary,
        "signals": signals,
        "assessment_suggestions": suggestions,
        "questions": questions,
        "limitations": limitations,
        "evidence_context": evidence_context,
        "evidence_count": len(evidence),
        "extracted_text_chars": len(corpus),
    }
